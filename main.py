# main file
# version 2.6.2 of OpenPyXL
# to install: >>> pip install openpyxl==2.6.2

import openpyxl
from datetime import date

def passthrough(pos_report, report, precip, temphigh, passnumber, date):
    print("\nStarting pass " + str(passnumber) + " ... ")

    # pos report sheet is active by default (just for testing).
    # POS report sheets will be merged into one workbook
    # Named 1-7
    # sheet_pos = pos_report[str(passnumber)] ** to be used
    # sheet_pos = pos_report.active # for a single pos_report (PROTOTYPING)

    sheet_pos = pos_report[str(passnumber)] # name pos_report/report sheets
                                            # the proper day of month
    sheet_report = report[str(passnumber)]

    pos_reportData = {'Gift Total': None, 'Food  (Department)': None,
                      'Green Fees  (Department)': None,
                      'Accessories  (Sub Department)': None,
                      'Bags  (Sub Department)': None,
                      'Balls  (Sub Department)': None,
                      'Gloves  (Sub Department)': None,
                      'Hard Goods  (Department)': None,
                      'Rewards Club  (Sub Department)': None,
                      'Units  (Sub Department)': None,
                      'Rentals  (Department)': None,
                      'Lessons  (Department)': None,
                      'Range  (Department)': None,
                      'Headwear  (Sub Department)': None}

    # populate pos_reportData dict w/ department values
    for i in range(8, sheet_pos.max_row):
        department = sheet_pos['A' + str(i)].value
        if department in pos_reportData:
            for j in range(i, sheet_pos.max_row):
                current = sheet_pos['A' + str(j)].value
                if current == '  Department Totals':
                    # set dict value at key [department]
                    pos_reportData[department] = sheet_pos['J' + str(j)].value
                    break

    # populate pos_reportData dict w/ sub department values
    for i in range(8, sheet_pos.max_row):
        # index B(i) is stored in sub_department
        sub_department = sheet_pos['B' + str(i)].value
        # if the variable is found in pos_reportData
        if sub_department in pos_reportData:
            for j in range(i, sheet_pos.max_row):
                current = sheet_pos['B' + str(j)].value
                if current == '  Sub Department Totals':
                    # store the net value in proper dict key
                    pos_reportData[sub_department] = sheet_pos[
                        'J' + str(j)].value
                    break

    # dict for report
    data = {'EMPLOYEE SALES': None, 'GIFT CERTIFICATES':
        pos_reportData['Gift Total'], 'DIRECT WHS': None,
            'FOOD AND DRINK': pos_reportData['Food  (Department)'],
            'GREEN FEES TOTAL': pos_reportData['Green Fees  (Department)'],
            'PASSES & MINI GOLF TOTAL': None,
            'ACCESSORIES': pos_reportData['Accessories  (Sub Department)'],
            'BAGS': pos_reportData['Bags  (Sub Department)'],
            'BALLS': pos_reportData['Balls  (Sub Department)'], 'CARTS': None,
            'GLOVES': pos_reportData['Gloves  (Sub Department)'],
            'JUNIOR CLUBS': None, "MEN'S CLUBS": None, "LADIES' CLUBS": None,
            "SHOES": None, 'LEAGUES': None, '9&DINE OTHER:': None,
            '9&DINE GF': None,
            'JUNIOR CLUB': None, 'MLGC OTHER': None, 'MLGC GF': None,
            'TOTAL LEAGUE (INCL GF)': None,
            'LESSONS': pos_reportData['Lessons  (Department)'],
            'MEMBERSHIPS': pos_reportData['Rewards Club  (Sub Department)'],
            'RANGE TOKENS': pos_reportData['Units  (Sub Department)'],
            'RENTALS': pos_reportData['Rentals  (Department)'],
            'HEADWEAR': pos_reportData['Headwear  (Sub Department)'],
            "LAST YEAR'S DATE": date, 'TEMPERATURE HIGH': temphigh,
            'PRECIPITATION': precip}

    for i in range(1, sheet_report.max_row):
        value = sheet_report['A' + str(i)].value
        if value in data:
            value = value.strip()
            # print('Match found: ' + str(value))
            sheet_report['C' + str(i)].value = data[value]
            # print('Addition successful!')

    for i in range(1, sheet_report.max_row):
        value = sheet_report['B' + str(i)].value
        if value in data:
            value = value.strip()
            # print('Match found: ' + str(value))
            sheet_report['C' + str(i)].value = data[value]
            # print('Addition successful!')

    print("Pass " + str(passnumber) + " done!\n")


today = date.today()

weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

print("Hi Ange, welcome to the automation wizard!")
print("You will be prompted for a filename")
print("Please include the file extension at the end (e.g. .xlsx) \n")
userResponse = input("Enter the filename of POS report: ")

wb_pos = openpyxl.load_workbook(userResponse)

userResponse = input("Enter the filename of report template: ")

wb_report = openpyxl.load_workbook(userResponse)

month = today.strftime("%b")

month_num = today.month

day = today.strftime("%d ")

year = today.year

start = int(input("Enter day to start at: "))

end = int(input("Enter last day: "))

dayofweek = today.weekday()

startDate = (end - (end - start))

ogDate = str(weekdays[int(dayofweek) - (end - start)]) + ", " + str(int(day) -
    (end - start)) + " " + month

for z in range(start, end):
    weekday = weekdays[int(dayofweek) - (end - z)]
    # day of the week, day, month
    date = str(weekday) + ", " + str(int(day) - (end - z)) + " " + month
    print(date)
    startDate += 1
    precip = input("Enter precipitation: ")
    tempHigh = input("Enter temperature high: ")
    passthrough(pos_report=wb_pos, report=wb_report, precip=precip,
                temphigh=tempHigh, passnumber=z, date=date)

wb_report.save("report_completed.xlsx")

print("Sheets from " + ogDate + " to " + date + " populated.")

input()