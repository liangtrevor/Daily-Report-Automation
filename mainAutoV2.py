# main file
# version 2.6.2 of OpenPyXL
# to install: >>> pip install openpyxl==2.6.2

import openpyxl
from datetime import date
# import pprint

today = date.today()
# days = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
days = {'0': 'Mon', '1': 'Tue', '2': 'Wed', '3': 'Thu', '4': 'Fri',
        '5': 'Sat', '6': 'Sun'}

print("Hi, welcome to the automation wizard!")
print("You will be prompted for filenames.")
print("Please include the file extension at the end (e.g. .xlsx) \n")
userResponse = input("Enter the filename of POS report: ")

wb_pos = openpyxl.load_workbook(userResponse)

sheet_pos = wb_pos.active

userResponse = input("Enter the filename of report template: ")

wb_report = openpyxl.load_workbook(userResponse)

# filename for saving new sheet at the end
newFilename = userResponse + "_completed"

userResponse = input("Enter the sheet name: ")

sheet_report = wb_report[userResponse]

userResponse = input("Enter year: ")

sheet_report['C' + str(1)].value = userResponse

userResponse = input("Enter date ('t' for today's date): ")

if userResponse != 't':
    date = userResponse
else:
    month = today.strftime("%b")

    day = today.strftime("%d ")

    daywk = days[str(today.weekday())]

    date = str(daywk) + ", " + day + month

userResponse = input("Enter temperature high: ")

tempHigh = userResponse

userResponse = input("Enter precipitation: ")

precip = userResponse

print("\nStarting ... \n")

pos_reportData = {'Gift Total': None, 'Food  (Department)': None,
                  'Green Fees  (Department)': None,
                  'Accessories  (Sub Department)': None,
                  'Bags  (Sub Department)': None,
                  'Balls  (Sub Department)': None,
                  'Gloves  (Sub Department)': None,
                  'Hard Goods  (Department)': None,
                  'Rewards Club  (Sub Department)': None,
                  'Units  (Sub Department)': None,
                  'Rentals  (Department)': None,
                  'Lessons  (Department)': None,
                  'Range  (Department)': None}

# populate pos_reportData dict w/ department values
for i in range(8, sheet_pos.max_row):
    department = sheet_pos['A' + str(i)].value
    if department in pos_reportData:
        for j in range(i, sheet_pos.max_row):
            current = sheet_pos['A' + str(j)].value
            if current == '  Department Totals':
                # set dict value at key [department]
                pos_reportData[department] = sheet_pos['J' + str(j)].value
                break

# populate pos_reportData dict w/ sub department values
for i in range(8, sheet_pos.max_row):
    # index B(i) is stored in sub_department
    sub_department = sheet_pos['B' + str(i)].value
    # if the variable is found in pos_reportData
    if sub_department in pos_reportData:
        for j in range(i, sheet_pos.max_row):
            current = sheet_pos['B' + str(j)].value
            if current == '  Sub Department Totals':
                # store the net value in proper dict key
                pos_reportData[sub_department] = sheet_pos['J' + str(j)].value
                break

# dict for report
data = {'EMPLOYEE SALES': None, 'GIFT CERTIFICATES':
    pos_reportData['Gift Total'], 'DIRECT WHS': None,
    'FOOD AND DRINK': pos_reportData['Food  (Department)'],
    'GREEN FEES TOTAL': pos_reportData['Green Fees  (Department)'],
    'PASSES & MINI GOLF TOTAL': None,
    'ACCESSORIES': pos_reportData['Accessories  (Sub Department)'],
    'BAGS': pos_reportData['Bags  (Sub Department)'],
    'BALLS': pos_reportData['Balls  (Sub Department)'], 'CARTS': None,
    'GLOVES': pos_reportData['Gloves  (Sub Department)'],
    'JUNIOR CLUBS': None, "MEN'S CLUBS": None, "LADIES' CLUBS": None,
    "SHOES": None, 'LEAGUES': None, '9&DINE OTHER:': None, '9&DINE GF': None,
    'JUNIOR CLUB': None, 'MLGC OTHER': None, 'MLGC GF': None,
    'TOTAL LEAGUE (INCL GF)': None,
    'LESSONS': pos_reportData['Lessons  (Department)'],
    'MEMBERSHIPS': pos_reportData['Rewards Club  (Sub Department)'],
    'RANGE TOKENS': pos_reportData['Units  (Sub Department)'],
    'RENTALS': pos_reportData['Rentals  (Department)'],
    "LAST YEAR'S DATE": date, 'TEMPERATURE HIGH': tempHigh,
    'PRECIPITATION': precip}

for i in range(1, sheet_report.max_row):
    value = sheet_report['A' + str(i)].value
    if value in data:
        value = value.strip()
        # print('Match found: ' + str(value))
        sheet_report['C' + str(i)].value = data[value]
        # print('Addition successful!')

for i in range(1, sheet_report.max_row):
    value = sheet_report['B' + str(i)].value
    if value in data:
        value = value.strip()
        # print('Match found: ' + str(value))
        sheet_report['C' + str(i)].value = data[value]
        # print('Addition successful!')

wb_report.save(newFilename + ".xlsx")

# Testing block

# print("Writing pos_reportData dict... (for internal use)")
# resultFile = open('pos_data_pprinted_main.py', 'w')
# resultFile.write('allData =' + pprint.pformat(pos_reportData))
# resultFile.close()
# print("Done.")
#
# print("Writing data dict... (for internal use)")
# resultFile = open('data_pprinted_main.py', 'w')
# resultFile.write('allData =' + pprint.pformat(data))
# resultFile.close()
# print("Done.")

print("Done!")

input()