# compartmentalize main - generalization
# version 2.6.2 of OpenPyXL
# to install: >>> pip install openpyxl==2.6.2

import openpyxl
from datetime import date
# import pprint

today = date.today()
# days = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
days = {'0': 'Mon', '1': 'Tue', '2': 'Wed', '3': 'Thu', '4': 'Fri',
        '5': 'Sat', '6': 'Sun'}

# one pass
def passthrough(pos_report, report, precip, tempHigh):
    for z in range(8):
        print("\nStarting pass " + str(z) + " ... \n")

        # pos report sheet is active by default.
        # TODO: create option for formatting POS report so multiple sheets are in one
        sheet_pos = pos_report.active
        sheet_report = report[str(z)]

        pos_reportData = {'Gift Total': None, 'Food  (Department)': None,
                          'Green Fees  (Department)': None,
                          'Accessories  (Sub Department)': None,
                          'Bags  (Sub Department)': None,
                          'Balls  (Sub Department)': None,
                          'Gloves  (Sub Department)': None,
                          'Hard Goods  (Department)': None,
                          'Rewards Club  (Sub Department)': None,
                          'Units  (Sub Department)': None,
                          'Rentals  (Department)': None,
                          'Lessons  (Department)': None,
                          'Range  (Department)': None}

        # populate pos_reportData dict w/ department values
        for i in range(8, sheet_pos.max_row):
            department = sheet_pos['A' + str(i)].value
            if department in pos_reportData:
                for j in range(i, sheet_pos.max_row):
                    current = sheet_pos['A' + str(j)].value
                    if current == '  Department Totals':
                        # set dict value at key [department]
                        pos_reportData[department] = sheet_pos['J' + str(j)].value
                        break

        # populate pos_reportData dict w/ sub department values
        for i in range(8, sheet_pos.max_row):
            # index B(i) is stored in sub_department
            sub_department = sheet_pos['B' + str(i)].value
            # if the variable is found in pos_reportData
            if sub_department in pos_reportData:
                for j in range(i, sheet_pos.max_row):
                    current = sheet_pos['B' + str(j)].value
                    if current == '  Sub Department Totals':
                        # store the net value in proper dict key
                        pos_reportData[sub_department] = sheet_pos[
                            'J' + str(j)].value
                        break

        # dict for report
        data = {'EMPLOYEE SALES': None, 'GIFT CERTIFICATES':
            pos_reportData['Gift Total'], 'DIRECT WHS': None,
                'FOOD AND DRINK': pos_reportData['Food  (Department)'],
                'GREEN FEES TOTAL': pos_reportData['Green Fees  (Department)'],
                'PASSES & MINI GOLF TOTAL': None,
                'ACCESSORIES': pos_reportData['Accessories  (Sub Department)'],
                'BAGS': pos_reportData['Bags  (Sub Department)'],
                'BALLS': pos_reportData['Balls  (Sub Department)'], 'CARTS': None,
                'GLOVES': pos_reportData['Gloves  (Sub Department)'],
                'JUNIOR CLUBS': None, "MEN'S CLUBS": None, "LADIES' CLUBS": None,
                "SHOES": None, 'LEAGUES': None, '9&DINE OTHER:': None,
                '9&DINE GF': None,
                'JUNIOR CLUB': None, 'MLGC OTHER': None, 'MLGC GF': None,
                'TOTAL LEAGUE (INCL GF)': None,
                'LESSONS': pos_reportData['Lessons  (Department)'],
                'MEMBERSHIPS': pos_reportData['Rewards Club  (Sub Department)'],
                'RANGE TOKENS': pos_reportData['Units  (Sub Department)'],
                'RENTALS': pos_reportData['Rentals  (Department)'],
                "LAST YEAR'S DATE": date, 'TEMPERATURE HIGH': tempHigh,
                'PRECIPITATION': precip}

        for i in range(1, sheet_report.max_row):
            value = sheet_report['A' + str(i)].value
            if value in data:
                value = value.strip()
                # print('Match found: ' + str(value))
                sheet_report['C' + str(i)].value = data[value]
                # print('Addition successful!')

        for i in range(1, sheet_report.max_row):
            value = sheet_report['B' + str(i)].value
            if value in data:
                value = value.strip()
                # print('Match found: ' + str(value))
                sheet_report['C' + str(i)].value = data[value]
                # print('Addition successful!')

        report.save("report_completed_" + str(z) + ".xlsx")


input()